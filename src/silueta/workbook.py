"""Messy-Excel recovery: turn real-world workbook sheets into profil-able tables.

Real vendor workbooks have title rows above the header, multi-row headers,
footer rows, blank padding, and values typed as text. This module extracts a
usable table region per sheet and loads it into DuckDB as all-VARCHAR data;
type recovery (TRY_CAST promotion) happens downstream in the profiler so
Excel and CSV share one code path.

Raw cell values live only inside this process; nothing here emits output.
"""

from __future__ import annotations

import datetime as dt
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import duckdb
from openpyxl import load_workbook

MAX_HEADER_SCAN_ROWS = 10
FOOTER_TOKENS = re.compile(r"^\s*(total|subtotal|grand total|sum|notes?|source)\b", re.IGNORECASE)


@dataclass
class SheetTable:
    """One recovered table region from one worksheet."""

    sheet_name: str
    columns: list[str]
    rows: list[tuple]  # all values stringified (or None)
    recovery: dict[str, Any] = field(default_factory=dict)
    not_tabular: bool = False


def _cell_str(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, dt.datetime):
        # openpyxl resolves Excel serial dates to datetimes; midnight -> date.
        if value.time() == dt.time(0, 0):
            return value.date().isoformat()
        return value.isoformat(sep=" ")
    if isinstance(value, dt.date):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = str(value).strip()
    return text if text != "" else None


def _fill(row: list[str | None]) -> int:
    return sum(1 for v in row if v is not None)


def _looks_headerish(row: list[str | None]) -> bool:
    """A header row is mostly non-numeric strings."""
    values = [v for v in row if v is not None]
    if len(values) < 2:
        return False
    numeric = sum(1 for v in values if _is_numberlike(v))
    return numeric / len(values) < 0.5


def _is_numberlike(value: str) -> bool:
    try:
        float(value.replace(",", ""))
        return True
    except ValueError:
        return False


def _find_header_row(grid: list[list[str | None]]) -> int | None:
    """Pick the header row: the best-filled header-ish row near the top whose
    following rows actually contain data."""
    best, best_score = None, 0.0
    scan = min(len(grid), MAX_HEADER_SCAN_ROWS)
    for i in range(scan):
        row = grid[i]
        if not _looks_headerish(row):
            continue
        data_below = sum(_fill(r) > 0 for r in grid[i + 1 : i + 6])
        if data_below == 0:
            continue
        score = _fill(row) + data_below * 0.1
        if score > best_score:
            best, best_score = i, score
    return best


def _merge_group_labels(
    parents: list[str | None], header: list[str | None]
) -> list[str | None]:
    """Merge a sparse group-label row above the header as 'parent child';
    parent labels forward-fill across their span."""
    merged: list[str | None] = []
    parent: str | None = None
    for top, bottom in zip(parents, header, strict=False):
        if top is not None:
            parent = top
        if bottom is None:
            merged.append(parent)
        elif parent is not None:
            merged.append(f"{parent} {bottom}")
        else:
            merged.append(bottom)
    return merged


def _dedupe_names(names: list[str | None]) -> list[str]:
    seen: dict[str, int] = {}
    out: list[str] = []
    for i, raw in enumerate(names):
        name = (raw or f"column_{i + 1}").strip()
        key = name.lower()
        if key in seen:
            seen[key] += 1
            name = f"{name}_{seen[key]}"
        else:
            seen[key] = 1
        out.append(name)
    return out


def extract_tables(path: Path) -> list[SheetTable]:
    if path.suffix.lower() == ".xls":
        raise NotImplementedError(
            "Legacy .xls is not supported — convert to .xlsx first "
            "(e.g. `libreoffice --convert-to xlsx`), then rerun."
        )
    wb = load_workbook(path, read_only=True, data_only=True)
    tables = []
    for ws in wb.worksheets:
        tables.append(_extract_sheet(ws))
    wb.close()
    return tables


def _extract_sheet(ws) -> SheetTable:
    grid = [[_cell_str(cell) for cell in row] for row in ws.iter_rows(values_only=True)]
    # Trim fully-empty rows at both ends and fully-empty trailing columns.
    while grid and _fill(grid[0]) == 0:
        grid.pop(0)
    while grid and _fill(grid[-1]) == 0:
        grid.pop()
    if grid:
        width = max(len(r) for r in grid)
        grid = [list(r) + [None] * (width - len(r)) for r in grid]
        used = [i for i in range(width) if any(r[i] is not None for r in grid)]
        grid = [[r[i] for i in used] for r in grid]

    recovery: dict[str, Any] = {}
    header_idx = _find_header_row(grid) if grid else None
    if header_idx is None:
        return SheetTable(ws.title, [], [], {"reason": "no header row found"}, not_tabular=True)

    header = grid[header_idx]
    skipped = header_idx
    # A sparse header-ish row immediately above the header is a group-label
    # row (e.g. "Provider" spanning NPI+Name) — merge it, don't skip it.
    if header_idx > 0:
        above = grid[header_idx - 1]
        # Require >=2 filled cells: a lone title cell is not a group-label row.
        if _fill(above) >= 2 and (_looks_headerish(above) or _fill(above) < _fill(header)):
            header = _merge_group_labels(above, header)
            recovery["multirow_header"] = True
            skipped = header_idx - 1
    if skipped:
        recovery["skipped_leading_rows"] = skipped
    data_start = header_idx + 1

    data = [tuple(r) for r in grid[data_start:] if _fill(r) > 0]

    # Drop footer rows: trailing rows that are sparse or start with total-ish tokens.
    footer_dropped = 0
    while data:
        last = data[-1]
        sparse = _fill(list(last)) <= max(1, len(header) // 3)
        totalish = last[0] is not None and FOOTER_TOKENS.match(str(last[0]))
        if sparse or totalish:
            data.pop()
            footer_dropped += 1
        else:
            break
    if footer_dropped:
        recovery["dropped_footer_rows"] = footer_dropped

    if len(header) < 2 or not data:
        return SheetTable(ws.title, [], [], {"reason": "no usable data region"}, not_tabular=True)

    return SheetTable(ws.title, _dedupe_names(header), data, recovery)


def register_table(conn: duckdb.DuckDBPyConnection, table: SheetTable, view_name: str) -> None:
    """Load a SheetTable into DuckDB as an all-VARCHAR table; the profiler's
    type recovery promotes columns afterward."""
    cols = ", ".join(f'"{c.replace(chr(34), chr(34) * 2)}" VARCHAR' for c in table.columns)
    conn.execute(f'CREATE OR REPLACE TEMP TABLE "{view_name}" ({cols})')
    placeholders = ", ".join("?" for _ in table.columns)
    conn.executemany(
        f'INSERT INTO "{view_name}" VALUES ({placeholders})',
        [row[: len(table.columns)] + (None,) * (len(table.columns) - len(row)) for row in table.rows],
    )

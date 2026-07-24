"""The ugly-workbook corpus: synthetic vendor-Excel horrors, no real data.

Exit test for messy-Excel recovery — every sheet here mimics a documented
real-world failure mode, and profiling must get headers and types right
without hand edits.
"""

from __future__ import annotations

import datetime as dt
import json
from pathlib import Path

import pytest
from openpyxl import Workbook

from silueta.profiler import profile_paths
from silueta.report import render_markdown
from silueta.workbook import extract_tables

SENTINEL_NAMES = [f"Vqzworker{i}" for i in range(40)]


@pytest.fixture()
def ugly_workbook(tmp_path: Path) -> Path:
    wb = Workbook()

    # Sheet 1: title rows above header, footer total row, numbers stored as text.
    ws = wb.active
    ws.title = "claims"
    ws.append(["Quarterly Claims Extract"])
    ws.append(["Prepared by vendor — CONFIDENTIAL"])
    ws.append([])
    ws.append(["Claim ID", "Member", "Amount", "Service Date"])
    for i in range(40):
        ws.append([f"CLM-{4000 + i}", SENTINEL_NAMES[i], f"1,2{i % 10:02d}.50", dt.datetime(2025, (i % 12) + 1, 3)])
    ws.append(["Total", None, "48,900.00", None])

    # Sheet 2: two-row header (group labels over fields), blank padding column.
    ws2 = wb.create_sheet("providers")
    ws2.append(["Provider", None, "Contact", None, None])
    ws2.append(["NPI", "Name", "Email", "Phone", None])
    for i in range(30):
        ws2.append([f"{1234567890 + i}", f"Vqzclinic {i}", f"vqz{i}@sentinel-med.org", f"555-01{i:02d}", None])

    # Sheet 3: not a table — a dashboard-ish note sheet.
    ws3 = wb.create_sheet("notes")
    ws3.append(["Remember to refresh the pivot"])
    ws3.append([])
    ws3.append(["Contact Bob"])

    path = tmp_path / "vendor.xlsx"
    wb.save(path)
    return path


def test_region_and_header_recovery(ugly_workbook: Path):
    tables = {t.sheet_name: t for t in extract_tables(ugly_workbook)}

    claims = tables["claims"]
    assert claims.columns == ["Claim ID", "Member", "Amount", "Service Date"]
    assert claims.recovery["skipped_leading_rows"] == 3
    assert claims.recovery["dropped_footer_rows"] == 1
    assert len(claims.rows) == 40

    providers = tables["providers"]
    assert providers.recovery.get("multirow_header") is True
    assert providers.columns == ["Provider NPI", "Provider Name", "Contact Email", "Contact Phone"]

    assert tables["notes"].not_tabular is True


def test_type_recovery(ugly_workbook: Path):
    profile = profile_paths([ugly_workbook])
    tables = {t["name"]: t for t in profile["tables"]}

    claims = {c["name"]: c for c in tables["vendor/claims"]["columns"]}
    assert claims["Amount"]["type_recovered"]["to"] == "DOUBLE"
    assert claims["Amount"]["physical_type"] == "DOUBLE"
    assert claims["Service Date"]["type_recovered"]["to"] in ("DATE", "TIMESTAMP")
    assert claims["Claim ID"]["physical_type"] == "VARCHAR"
    assert claims["Claim ID"]["masks"][0]["mask"] == "AAA-9999"

    providers = {c["name"]: c for c in tables["vendor/providers"]["columns"]}
    assert providers["Provider NPI"]["type_recovered"]["to"] == "BIGINT"
    assert any(h["type"] == "email" for h in providers["Contact Email"]["semantic"])

    assert tables["vendor/notes"]["not_tabular"] is True


def test_workbook_contract(ugly_workbook: Path):
    profile = profile_paths([ugly_workbook])
    outputs = json.dumps(profile) + render_markdown(profile)
    for fragment in ("Vqzworker", "Vqzclinic", "sentinel-med", "CLM-40"):
        assert fragment not in outputs, f"workbook value leaked: {fragment}"
    # Title rows are dropped, not echoed.
    assert "CONFIDENTIAL" not in outputs
